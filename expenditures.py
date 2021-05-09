from openpyxl import *
from openpyxl.styles import *
from datetime import date
from time import sleep
import os

def CreateOrOpenTable():
    '''Creates an excel file and types formatted titles or opens an existing one'''
    try:
        book=load_workbook("expenses.xlsx")
        book.save("expenses.xlsx")
    except:
        book = Workbook()
        sheet=book.active
        sheet["A1"],sheet["B1"],sheet["C1"],sheet["D1"]=\
            "Num","Purpose","Cost","Date"
        for cell in ["A1","B1","C1","D1"]:
            sheet[cell].font=Font(bold=True)
            sheet[cell].alignment=Alignment(horizontal="center",vertical="center")
        book.save("expenses.xlsx")
    return book

def appendTable(lst):
    book=load_workbook('expenses.xlsx')
    sheet=book.active
    if sheet.max_row==1:
        i=0
    else:
        i=sheet.max_row-1
    for line in lst:
        i+=1
        line=[i]+line[:]
        sheet.append(line)
    book.save('expenses.xlsx')

def checkDate(Date):
    d,m,y=int(Date[0]),int(Date[1]),int(Date[2])
    try:
        return date(y,m,d)<=date.today()
    except:
        return False    
def printTable():
    book=load_workbook('expenses.xlsx')
    sheet=book.active
    max_col=int(sheet.max_column)
    max_row=int(sheet.max_row)
    print("\n")
    for Row in range(1,max_row+1):
        for Col in range(1,max_col+1):
            print("{:20s}".format(str(sheet.cell(row=Row,column=Col).value)),end="")
        print()
    book.save("expenses.xlsx")
    
def main():
    global book
    print("[LIST OF EXPENCES]\n")
    book=CreateOrOpenTable()
    info=[]
    while True:
        purpose=input("Type a purpose/reason for expenditure:").strip()
        while True:
            cost=input("Type a cost of it:(in dollars)")
            try:
                cost=float(cost)
                break
            except:
                temp=0
        while True:
            Date=input("Type date of it:(e.g DD.MM.YYYY)").split(".")
            if checkDate(Date):
                    break
        info.append([purpose,cost,str(date(int(Date[2]),int(Date[1]),int(Date[0])))[:10]])
        check=input("Do you want to add something?:(split line if NO!)")
        if len(check)==0:
            sleep(1.2)
            os.system("cls")
            break
    appendTable(info)
    printTable()
    sleep(10000)
    
if __name__=="__main__":
    main()
